# -*- coding: utf-8 -*-

from io import BytesIO
from typing import Any, Dict, Iterator, List, Optional, Tuple

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class SuchIdIsExistException(Exception):
    """Raised when the document with the identifier is already in the cache."""
    pass


class NoSuchIdException(Exception):
    """Raised when accessing an absent document identifier."""
    pass


class NoOpenedDocumentsException(Exception):
    """Raised in the absence of open documents."""
    pass


class ExcelLibrary(object):
    """Library for working with Excel documents.

    == Dependencies ==
    | robot framework | http://robotframework.org |

    == Example ==
    | *Settings* | *Value* |
    | Library    | ExcelLibrary.py |
    | Library    | Collections |

    | *Test Cases* | *Action* | *Argument* | *Argument* | *Argument* |
    | Simple |
    |    | Create Excel Document | doc_id=docname1 |
    |    | Write Excel Cell | row_num=1 | col_num=1 | value=text |
    |    | Save Excel Document | filename=file.xlsx |
    |    | Close Current Excel Document |
    """

    ROBOT_LIBRARY_SCOPE = 'GLOBAL'

    def __init__(self) -> None:
        """Initializer"""
        self._cache: Dict[str, openpyxl.Workbook] = {}
        self._current_id: Optional[str] = None

    def create_excel_document(self, doc_id: str) -> str:
        """Creates new excel document.\n
        *Args:*\n
            _doc_id_: document identifier in the cache.\n
        *Returns:*\n
            Identifier of created document.\n
        *Example:*\n
        | Create Excel Document | doc_id=doc |
        | Close All Excel Documents |
        """
        doc_id = str(doc_id)
        if doc_id in self._cache:
            message = u"Document with such id {0} is created."
            raise SuchIdIsExistException(message.format(doc_id))
        workbook = openpyxl.Workbook()
        self._cache[doc_id] = workbook
        self._current_id = doc_id
        return self._current_id

    def open_excel_document(self, filename: str, doc_id: str) -> str:
        """Opens xlsx document file.\n
        *Args:*\n
            _filename_: document name.\n
            _doc_id_: the identifier for the document that will be opened.\n
        *Returns:*\n
            Document identifier from the cache.\n
        *Example:*\n
        | Open Excel Document | filename=file.xlsx | doc_id=docid |
        | Close All Excel Documents |
        """
        filename = str(filename)
        doc_id = str(doc_id)
        if doc_id in self._cache:
            message = u"Document with such id {0} is opened."
            raise SuchIdIsExistException(message.format(doc_id))
        workbook = openpyxl.load_workbook(filename=filename)
        self._cache[doc_id] = workbook
        self._current_id = doc_id
        return self._current_id

    def open_excel_document_from_stream(self, stream: bytes, doc_id: str) -> str:
        """Opens xlsx document from stream.\n
        *Args:*\n
            _stream_: file-like byte stream object {e.g. from any http request).\n
            _doc_id_: the identifier for the document that will be opened.\n
        *Returns:*\n
            Document identifier from the cache.\n
        *Example:*\n
        | Open Excel Document From Stream | stream=${report} | doc_id=report.xlsx |
        | Close All Excel Documents |
        """
        doc_id = str(doc_id)
        if doc_id in self._cache:
            message = u"Document with such id {0} is opened."
            raise SuchIdIsExistException(message.format(doc_id))
        workbook = openpyxl.load_workbook(filename=BytesIO(stream))
        self._cache[doc_id] = workbook
        self._current_id = doc_id
        return self._current_id

    def make_list_from_excel_sheet(self, sheet: Worksheet) -> list:
        """Making list from Excel sheet.\n
        *Args:*\n
            _sheet_: Excel file sheet.\n
        *Returns:*\n
            _data_: a list of tuples corresponding to the values of each line of an Excel file.
        *Example:*\n
        | ${excel_data_list}= | Make List From Excel Sheet | sheet1 |
        """
        data = []
        for row in sheet.values:
            data.append(row)
        return data

    def switch_current_excel_document(self, doc_id: str) -> Optional[str]:
        """Switches current excel document.\n
        *Args:*\n
            _doc_id_: identifier of the document to switch.\n
        *Returns:*\n
            Identifier of previous document.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | docname1 |
        | ${doc2}= | Create Excel Document | docname2 |
        | Switch Current Excel Document | ${doc1} |
        | Close All Excel Documents |
        """
        if doc_id not in self._cache:
            message = u"Document with such id {0} is not opened yet."
            raise NoSuchIdException(message.format(doc_id))
        old_name = self._current_id
        self._current_id = doc_id
        return old_name

    def close_current_excel_document(self) -> Optional[str]:
        """Closes current document.\n
        *Returns:*\n
            Closed document identifier.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | docname1 |
        | ${doc2}= | Create Excel Document | docname2 |
        | Close Current Excel Document |
        """
        if self._current_id is not None:
            self._cache.pop(self._current_id)
            self._current_id = None
        if self._cache:
            self._current_id = list(self._cache.keys())[0]
        return self._current_id

    def close_all_excel_documents(self) -> None:
        """Closes all opened documents.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | docname1 |
        | ${doc2}= | Create Excel Document | docname2 |
        | Close All Excel Documents |
        """
        self._cache = {}
        self._current_id = None

    def save_excel_document(self, filename: str) -> None:
        """Saves the current document to disk.\n
        *Args:*\n
            _filename_: file name to save.\n
        *Example:*\n
        | Create Excel Document | doc_id=doc1 |
        | Save Excel Document | filename=file1.xlsx |
        | Close All Excel Documents |
        """
        workbook = self._get_current_workbook()
        workbook.save(filename=filename)

    def get_list_sheet_names(self) -> List[str]:
        """Returns a list of sheet names in the current document.\n
        *Returns:*\n
            List of page names.\n
        *Example:*\n
        | Create Excel Document | doc_id=doc1 |
        | ${sheets}= | Get List Sheet Names |
        | List Should Contain Value | ${sheets} | sheet1 |
        | Close All Excel Documents |
        """
        workbook = self._get_current_workbook()
        return workbook.sheetnames

    def _get_current_workbook(self) -> openpyxl.Workbook:
        """Checks opened document.\n
        *Returns:*\n
            Current document.\n
        """
        if not self._cache or not self._current_id:
            raise NoOpenedDocumentsException(u"No opened documents in cache.")
        return self._cache[self._current_id]

    def get_sheet(self, sheet_name: str = None) -> Worksheet:
        """Returns a page from the current document.\n
        *Args:*\n
            _sheet_name_: sheet name.\n
        *Returns:*\n
            Document's sheet.\n
        """
        workbook = self._get_current_workbook()
        if sheet_name is None:
            return workbook.active
        sheet_name = str(sheet_name)
        return workbook[sheet_name]

    def read_excel_cell(self, row_num: int, col_num: int, sheet_name: str = None) -> Any:
        """Returns content of a cell.\n
        *Args:*\n
            _row_num_: row number, starts with 1.\n
            _col_num_: column number, starts with 1.\n
            _sheet_name_: sheet name, where placed cell, that need to be read.\n
        *Returns:*\n
            Content of the cell in the specified column and row.\n
        *Example:*\n
        | Open Excel Document | filename=file1.xlsx | doc_id=doc1 |
        | ${a1}= | Read Excel Cell | row_num=1 | col_num=1 |
        | Should Be Equal As Strings | ${a1} | text |
        | Close All Excel Documents |
        """
        row_num = int(row_num)
        col_num = int(col_num)
        sheet = self.get_sheet(sheet_name)
        cell: Cell = sheet.cell(row=row_num, column=col_num)
        return cell.value

    def read_excel_row(self, row_num: int, col_offset: int = 0, max_num: int = 0, sheet_name: str = None) -> List[Any]:
        """Returns content of a row from the current sheet of the document.\n
        *Args:*\n
            _row_num_: row number, starts with 1.\n
            _col_offset_: column indent.\n
            _max_num_: maximum number of columns to read.\n
            _sheet_name_: sheet name, where placed row, that need to be read.\n
        *Returns:*\n
            List, that stores the contents of a row.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | doc_id=docname1 |
        | ${row_data}= | Create List | t1 | t2 | t3 |
        | Write Excel Row | row_num=5 | row_data=${row_data} | sheet_name=${DEFAULT_SHEET_NAME} |
        | ${rd}= | Read Excel Row  | row_num=5 | max_num=3 | sheet_name=${DEFAULT_SHEET_NAME} |
        | Lists Should Be Equal | ${row_data} | ${rd} |
        | Close All Excel Documents |
        """
        row_num = int(row_num)
        col_offset = int(col_offset)
        max_num = int(max_num)
        sheet = self.get_sheet(sheet_name)
        row_iter: Iterator[Tuple[Cell]] = sheet.iter_rows(min_row=row_num, max_row=row_num,
                                                          min_col=1 + col_offset,
                                                          max_col=col_offset + max_num)
        row: Tuple[Cell] = next(row_iter)
        return [cell.value for cell in row]

    def read_excel_column(self, col_num: int, row_offset: int = 0, max_num: int = 0,
                          sheet_name: str = None) -> List[Any]:
        """Returns content of a column from the current sheet of the document.\n
        *Args:*\n
            _col_num_: column number, starts with 1.\n
            _row_offset_: row indent.\n
            _max_num_: maximum number of rows to read.\n
            _sheet_name_: sheet name, where placed column,
            that need to be read.\n
        *Returns:*\n
            List, that stores the contents of a row.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | doc_id=docname1 |
        | ${col_data}= | Create List | a1 | a2 | a3 |
        | Write Excel Column | col_num=3 | col_data=${col_data} | sheet_name=${DEFAULT_SHEET_NAME} |
        | ${cd}= | Read Excel Column | col_num=3 | max_num=3 | sheet_name=${DEFAULT_SHEET_NAME}|
        | Lists Should Be Equal | ${col_data} | ${cd} |
        | Close All Excel Documents |
        """
        col_num = int(col_num)
        row_offset = int(row_offset)
        max_num = int(max_num)
        sheet = self.get_sheet(sheet_name)
        row_iter: Iterator[Tuple[Cell]] = sheet.iter_rows(min_col=col_num, max_col=col_num,
                                                          min_row=1 + row_offset,
                                                          max_row=row_offset + max_num)
        return [row[0].value for row in row_iter]

    def write_excel_cell(self, row_num: int, col_num: int, value: Any, sheet_name: str = None) -> None:
        """Writes value to the cell.\n
        *Args:*\n
            _row_num_: row number, starts with 1.\n
            _col_num_: column number, starts with 1.\n
            _value_: value for writing to a cell.\n
            _sheet_name_: sheet name for write.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | doc_id=docname1 |
        | Write Excel Cell | row_num=1 | col_num=3 | value=a3 | sheet_name=${DEFAULT_SHEET_NAME} |
        | Close All Excel Documents |
        """
        row_num = int(row_num)
        col_num = int(col_num)
        sheet = self.get_sheet(sheet_name)
        sheet.cell(row=row_num, column=col_num, value=value)

    def write_excel_row(self, row_num: int, row_data: List[Any], col_offset: int = 0, sheet_name: str = None) -> None:
        """Writes a row to the document.\n
        *Args:*\n
            _row_num_: row number, starts with 1.\n
            _row_data_: list of values for writing.\n
            _col_offset_: number of indent columns from start.\n
            _sheet_name_: sheet name for write.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | doc_id=docname1 |
        | ${row_data}= | Create List | a1 | a2 | a3 |
        | Write Excel Row | row_num=1 | row_data=${row_data} | sheet_name=${DEFAULT_SHEET_NAME} |
        | Close All Excel Documents |
        """
        row_num = int(row_num)
        col_offset = int(col_offset)
        sheet = self.get_sheet(sheet_name)
        for col_num in range(len(row_data)):
            sheet.cell(row=row_num, column=col_num + col_offset + 1, value=row_data[col_num])

    def write_excel_rows(self, rows_data: List[List[Any]], rows_offset: int = 0, col_offset: int = 0,
                         sheet_name: str = None) -> None:
        """Writes a list of rows to the document.\n
        *Args:*\n
            _rows_data_: list of rows for writing.\n
            _rows_offset_: number of indent rows from start.\n
            _col_offset_: number of indent columns from start.\n
            _sheet_name_: sheet name for write.\n
        """
        for row_num, row_data in enumerate(rows_data):
            self.write_excel_row(row_num + int(rows_offset) + 1, row_data, col_offset, sheet_name)

    def write_excel_column(self, col_num: int, col_data: List[Any], row_offset: int = 0,
                           sheet_name: str = None) -> None:
        """Writes the data to a column.\n
        *Args:*\n
            _col_num_: column number, starts with 1.\n
            _col_data_: list of values for writing.\n
            _row_offset_: number of indent rows from start.\n
            _sheet_name_: sheet name for write.\n
        *Example:*\n
        | ${doc1}= | Create Excel Document | doc_id=docname1 |
        | ${col_data}= | Create List | a1 | a2 | a3 |
        | Write Excel Column | col_num=1 | col_data=${col_data} | sheet_name=${DEFAULT_SHEET_NAME} |
        | Close All Excel Documents |
        """
        col_num = int(col_num)
        row_offset = int(row_offset)
        sheet = self.get_sheet(sheet_name)
        for row_num in range(len(col_data)):
            sheet.cell(column=col_num, row=row_num + row_offset + 1, value=col_data[row_num])
